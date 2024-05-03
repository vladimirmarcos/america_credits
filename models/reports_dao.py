import os
from .conexion_db import ConexionDB
import sqlite3
from docxtpl import DocxTemplate
from tkinter import messagebox
from openpyxl import Workbook
import ast
from .account_dao import search_data_account
from .credit_dao import search_account_credits_info,calculate_rest_of_the_credits_info,search_account, interest_calculation,search_mora
#from Procesamiento import procesar_dato_fecha
import processes
from processes.str_and_date_processes import processes_data_date,process_data_str_to_date


def create_reports_credits(fecha1,fecha2):
    conexion=ConexionDB()
    sql=f""" SELECT credito,cuenta,fecha,monto_financiado,cuotas from creditos WHERE fecha>={fecha1} and fecha<={fecha2} and estado=1"""
    conexion.cursor.execute(sql)
    informacion=conexion.cursor.fetchall()
    conexion.close()
    t=0
    if informacion:
        for i in informacion:
            inf=list(i)
            cuenta=search_data_account(inf[1])
            nombre=cuenta[0]
            inf.insert(2,nombre)
            inf[3]=processes_data_date(inf[3])
            informacion[t]=inf
            t=t+1
        generate_informs(informacion,fecha1,fecha2)
    else:
        messagebox.showerror("Error al generar informe","No se encontraron creditos emitidos en las fechas ingresadas")



def generate_informs(information,fecha1,fecha2):
    template_location=os.path.join(os.path.abspath(os.getcwd()),"templates/reports_templates/")
    archive_location=os.path.join(template_location, 'credits.docx')  
    white_model= DocxTemplate(archive_location)
    context={
         'informacion':information,
          'fecha1':processes_data_date(fecha1),
          'fecha2':processes_data_date(fecha2)
     }
    
    white_model.render(context)
    rute_save=os.path.join(os.path.abspath(os.getcwd()),"reportes/creditos_emitidos")
    archive="Informe_creditos_emitidos_entre"+f"{processes_data_date(fecha1)}"+"_y_"+f"{processes_data_date(fecha2)}"+".doc"
    archive_name=os.path.join(rute_save,archive)
    white_model.save(archive_name)
    respuesta=messagebox.askyesno("Generar informe","¿Desea imprimir el informe?")
    if respuesta:
        try:
             os.startfile(archive_name,"print")
        except:
            messagebox.showerror("No se pudo imprimir el archivo", f"No se pudo impirmir el archivo, \n la direccion del mismo es {archive_name}")
    else:
        messagebox.showinfo("No se pudo imprimir el archivo", f"No se imprime el archivo, \n la direccion del mismo es {archive_name}")

def create_reports_credits_excel(fecha1,fecha2):
    conexion=ConexionDB()
    sql=f""" SELECT credito,cuenta,fecha,monto_financiado,cuotas from creditos WHERE fecha>={fecha1} and fecha<={fecha2} and estado=1"""
    conexion.cursor.execute(sql)
    informacion=conexion.cursor.fetchall()
    conexion.close()
    t=0
    if informacion:
        for i in informacion:
            inf=list(i)
            cuenta=search_data_account(inf[1])
            nombre=cuenta[0]
            inf.insert(2,nombre)
            inf[3]=processes_data_date(inf[3])
            informacion[t]=inf
            t=t+1
        wb = Workbook()
        ws = wb.active
        ws.append(["credito","cuenta","nombre","fecha","monto_financiado","cuotas"])
        for i in informacion:
            ws.append([f"{int(i[0])}",f"{int(i[1])}",f"{i[2]}",f"{i[3]}",f"{float(i[4])}",f"{int(i[5])}"])
        rute_save=os.path.join(os.path.abspath(os.getcwd()),"reportes/creditos_emitidos")
        archive="Informe_creditos_emitidos_entre"+f"{processes_data_date(fecha1)}"+"_y_"+f"{processes_data_date(fecha2)}"+".xlsx"
        archive_name=os.path.join(rute_save,archive)
        wb.save(archive_name)
        messagebox.showinfo("Se genero el reporte", f"No se imprime el archivo, \n la direccion del mismo es {archive_name}")
        
    else:
        messagebox.showerror("Error al generar informe","No se encontraron creditos emitidos en las fechas ingresadas")



def create_reports_pay(fecha1,fecha2):
    conexion=ConexionDB()
    sql=f""" SELECT id_pagos,fecha_pago,datos,cuenta from pagos WHERE fecha_pago>={fecha1} and fecha_pago<={fecha2}"""
    conexion.cursor.execute(sql)
    informacion=conexion.cursor.fetchall()
    conexion.close()
    total=0
    p=0
    if informacion:
     for i in informacion:
        inf=list(i)
        inf[1]=processes_data_date(inf[1])
        diccionary = ast.literal_eval(inf[2])
        keys = list(diccionary.keys())
        auxiliary=search_data_account(i[3])
        inf.append(auxiliary[0])
        plus=0
        for i in keys:
            list_information=diccionary[i]
            for t in list_information:
                plus+=t[5]
        inf.append(plus)
        informacion[p]=inf
        p=p+1
        total+=plus
     template_location=os.path.join(os.path.abspath(os.getcwd()),"templates/reports_templates/")
     archive_location=os.path.join(template_location, 'pay.docx')  
     white_model= DocxTemplate(archive_location)
     context={
         'informacion':informacion,
          'fecha1':processes_data_date(fecha1),
          'fecha2':processes_data_date(fecha2),
          'total':total
     }
     white_model.render(context)
     rute_save=os.path.join(os.path.abspath(os.getcwd()),"reportes/recibos_emitidos")
     archive="reportes_recibos_emitidos_entre"+f"{processes_data_date(fecha1)}"+"_y_"+f"{processes_data_date(fecha2)}"+".doc"
     archive_name=os.path.join(rute_save,archive)
     white_model.save(archive_name)
     respuesta=messagebox.askyesno("Generar informe","¿Desea imprimir el informe?")
     if respuesta:
        try:
             os.startfile(archive_name,"print")
        except:
            messagebox.showerror("No se pudo imprimir el archivo", f"No se pudo impirmir el archivo, \n la direccion del mismo es {archive_name}")
     else:
        messagebox.showinfo("No se pudo imprimir el archivo", f" \n la direccion del mismo es {archive_name}")
    else:
         messagebox.showerror("Error al generar reporte","No se encontraron recibos emitidos")
    
def create_reports_pay_excel(fecha1,fecha2):
    conexion=ConexionDB()
    sql=f""" SELECT id_pagos,fecha_pago,datos,cuenta from pagos WHERE fecha_pago>={fecha1} and fecha_pago<={fecha2}"""
    conexion.cursor.execute(sql)
    informacion=conexion.cursor.fetchall()
    conexion.close()
    total=0
    p=0
    if informacion:
     for i in informacion:
        inf=list(i)
        inf[1]=processes_data_date(inf[1])
        diccionary = ast.literal_eval(inf[2])
        keys = list(diccionary.keys())
        auxiliary=search_data_account(i[3])
        inf.append(auxiliary[0])
        plus=0
        for i in keys:
            list_information=diccionary[i]
            for t in list_information:
                plus+=t[5]
        inf.append(plus)
        informacion[p]=inf
        p=p+1
        total+=plus
     wb = Workbook()
     ws = wb.active
     ws.append(["recibo","cuenta","nombre","Importe","fecha"])
     for i in informacion:
            ws.append([f"{int(i[0])}",f"{int(i[3])}",f"{i[4]}",f"{float(i[5])}"])
     ws.append(["Total","","",f"{total}"])
     rute_save=os.path.join(os.path.abspath(os.getcwd()),"reportes/recibos_emitidos")
     archive="Informe_recibos_emitidos_entre"+f"{processes_data_date(fecha1)}"+"_y_"+f"{processes_data_date(fecha2)}"+".xlsx"
     archive_name=os.path.join(rute_save,archive)
     wb.save(archive_name)
     messagebox.showinfo("Se Genero el Reporte", f" \n la direccion del mismo es {archive_name}")
    else:
        messagebox.showerror("Error al generar reporte","No se encontraron recibos emitidos")

def create_reports_rest_credit(fecha1,fecha):
    conexion=ConexionDB()
    sql=f""" SELECT cuenta,credito from creditos WHERE fecha<={fecha1} and estado='1'"""
    conexion.cursor.execute(sql)
    informacion=conexion.cursor.fetchall()
    conexion.close()
    t=0
    if informacion:
     for i in informacion:
        i=list(i)
        informacion[t]=i
        t+=1
     lista_de_listas_ordenada = sorted(informacion, key=lambda x: x[0])
     account=[]
     account.append(lista_de_listas_ordenada[0][0])
     t=0
     for i in lista_de_listas_ordenada:
        if i[0]==account[t]:
            pass
        else:
            account.append(i[0])
            t+=1
     t=0
     
     for i in account:
         list_credits=search_account_credits_info(i)
         account[t]=[account[t],list_credits]
         
         t+=1
     t=0
     total=0
     for i in account:
         rest_credit_list=calculate_rest_of_the_credits_info(i[1],fecha)
         total=total+rest_credit_list
         i.append(rest_credit_list)
         auxiliary=search_data_account(i[0])
         i.append(auxiliary[0])
         account[t]=i
         t+=1
     
     template_location=os.path.join(os.path.abspath(os.getcwd()),"templates/reports_templates/")
     archive_location=os.path.join(template_location, 'totales.docx')  
     white_model= DocxTemplate(archive_location)  
     context={
         'informacion':account,
          'hoy':processes_data_date(fecha1),
          'total':total
     }
     white_model.render(context)
     rute_save=os.path.join(os.path.abspath(os.getcwd()),"reportes/reportes_saldos_totales")
     archive="reportes_saldos_totales_hasta_entre"+f"{processes_data_date(fecha1)}"+".doc"
     archive_name=os.path.join(rute_save,archive)
     white_model.save(archive_name)
     respuesta=messagebox.askyesno("Generar informe","¿Desea imprimir el informe?")
     if respuesta:
        try:
             os.startfile(archive_name,"print")
        except:
            messagebox.showerror("No se pudo imprimir el archivo", f"No se pudo impirmir el archivo, \n la direccion del mismo es {archive_name}")
     else:
        messagebox.showinfo("No se pudo imprimir el archivo", f" \n la direccion del mismo es {archive_name}")
    else:
       messagebox.showerror("Error al generar reporte","No se encontraron saldos en la fecha especificada") 


def create_reports_rest_credit_excel(fecha1,fecha):
    conexion=ConexionDB()
    sql=f""" SELECT cuenta,credito from creditos WHERE fecha<={fecha1} and estado='1'"""
    conexion.cursor.execute(sql)
    informacion=conexion.cursor.fetchall()
    conexion.close()
    t=0
    if informacion:
     for i in informacion:
        i=list(i)
        informacion[t]=i
        t+=1
     lista_de_listas_ordenada = sorted(informacion, key=lambda x: x[0])
     account=[]
     account.append(lista_de_listas_ordenada[0][0])
     t=0
     for i in lista_de_listas_ordenada:
        if i[0]==account[t]:
            pass
        else:
            account.append(i[0])
            t+=1
     t=0
     
     for i in account:
         list_credits=search_account_credits_info(i)
         account[t]=[account[t],list_credits]
         
         t+=1
     t=0
     total=0
     for i in account:
         rest_credit_list=calculate_rest_of_the_credits_info(i[1],fecha)
         total=total+rest_credit_list
         i.append(rest_credit_list)
         auxiliary=search_data_account(i[0])
         i.append(auxiliary[0])
         account[t]=i
         t+=1
    
     wb = Workbook()
     ws = wb.active
     ws.append(["cuenta","nombre","Total"])
     for i in account:
            ws.append([f"{int(i[0])}",f"{(i[3])}",f"{i[2]}"])
     ws.append(["Total","",f"{total}"])
     rute_save=os.path.join(os.path.abspath(os.getcwd()),"reportes/reportes_saldos_totales")
     archive="reportes_saldos_totales_hasta_entre"+f"{processes_data_date(fecha1)}"+".xlsx"
     archive_name=os.path.join(rute_save,archive)
     wb.save(archive_name)
     messagebox.showinfo("Se Genero el Reporte", f" \n la direccion del mismo es {archive_name}")
    else:
        messagebox.showerror("Error al generar reporte","No se encontraron totales a la feche especificada")


def create_reports_rest_credit(fecha1,fecha):
    conexion=ConexionDB()
    sql=f""" SELECT credito,fecha_id,fecha,monto,acuenta,cuota from fechas_pagos WHERE fecha<{fecha1} and estado='Por Pagar'"""
    conexion.cursor.execute(sql)
    informacion=conexion.cursor.fetchall()
    conexion.close()
    if informacion:
        t=0
        for i in informacion:
         i=list(i)
         informacion[t]=i
         t+=1
        lista_de_listas_ordenada = sorted(informacion, key=lambda x: x[0])
        credit_list=[]
        auxiliary=[]
        for i in lista_de_listas_ordenada:
            auxiliary.append(i[0])
        credit_list=set(auxiliary)
        
        dictionary_mora={}
        for elemento in credit_list:
            dictionary_mora[search_account(elemento)]=[]
        
        auxiliary=[]
        total_mora=0
        for elemento in credit_list:
            
            auxiliary=[]
            auxiliary.append(elemento)

            for i in lista_de_listas_ordenada:
                if i[0]==elemento:
                 date=process_data_str_to_date(i[2])
                 date=date.date()
                 interest=interest_calculation(fecha,date,i[3],search_mora())
                 total=i[3]+interest-i[4]
                 auxiliary.append([i[1],processes_data_date(i[2]),total,i[5]])
                 total_mora=total_mora+total
            auxiliary_end=dictionary_mora[search_account(elemento)]
            auxiliary_end.append(auxiliary)
            dictionary_mora[search_account(elemento)]=auxiliary_end
            
        
        claves_lista = list(dictionary_mora.keys())
        template_location=os.path.join(os.path.abspath(os.getcwd()),"templates/reports_templates/")
        archive_location=os.path.join(template_location, 'morosos.docx')  
        white_model= DocxTemplate(archive_location)
        context={
         'keys':claves_lista,
         'diccionario':dictionary_mora,
          'hoy':processes_data_date(fecha1)
            }
        white_model.render(context)
        rute_save=os.path.join(os.path.abspath(os.getcwd()),"reportes/reportes_morosos")
        archive="reportes_mororsos_a_la_fecha_"+f"{processes_data_date(fecha1)}"+".doc"
        archive_name=os.path.join(rute_save,archive)
        white_model.save(archive_name)
        respuesta=messagebox.askyesno("Generar informe","¿Desea imprimir el informe?")
        if respuesta:
         try:
             os.startfile(archive_name,"print")
         except:
            messagebox.showerror("No se pudo imprimir el archivo", f"No se pudo impirmir el archivo, \n la direccion del mismo es {archive_name}")
        else:
         messagebox.showinfo("No se pudo imprimir el archivo", f" \n la direccion del mismo es {archive_name}")            
                  

    else:
     messagebox.showerror("Error al generar reporte","No se encontraron morosos en la fecha especificada") 